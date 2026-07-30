"""Database seeder — parses the Aztec Excel dataset and populates the DB on startup."""

import asyncio
import logging
from datetime import datetime
import openpyxl

from sqlalchemy.ext.asyncio import AsyncSession

from app.domain.entities import Project, Task, TeamMember
from app.infrastructure.database import async_session_factory
from app.infrastructure.repositories.project_repository import ProjectRepository
from app.infrastructure.repositories.task_repository import TaskRepository
from app.infrastructure.repositories.team_repository import TeamRepository
from app.application.risk_engine import RiskEngineService
from app.infrastructure.auth import hash_password
from app.infrastructure.models import UserModel
from sqlalchemy import select

logger = logging.getLogger(__name__)

EXCEL_PATH = "/app/app/seed_data/dataset.xlsx"


def parse_date(date_str):
    if not date_str:
        return None
    if isinstance(date_str, datetime):
        return date_str.date()
    try:
        return datetime.strptime(str(date_str).split()[0], "%Y-%m-%d").date()
    except ValueError:
        return None


async def ensure_admin_user(session: AsyncSession):
    """Ensure the default user exists."""
    result = await session.execute(select(UserModel).where(UserModel.username == "projectuser"))
    if not result.scalar_one_or_none():
        user = UserModel(
            username="projectuser",
            hashed_password=hash_password("password123")
        )
        session.add(user)
        await session.flush()


async def seed_db():
    logger.info("Starting database seed process...")
    
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except FileNotFoundError:
        logger.warning(f"Seed dataset not found at {EXCEL_PATH}. Skipping seed.")
        return

    async with async_session_factory() as session:
        # Check if DB is already seeded (by checking if any team members exist)
        team_repo = TeamRepository(session)
        project_repo = ProjectRepository(session)
        task_repo = TaskRepository(session)
        
        existing = await team_repo.get_all(limit=1)
        if existing:
            logger.info("Database already seeded. Skipping.")
            return

        await ensure_admin_user(session)

        # 1. Seed Team Members
        if "Team" in wb.sheetnames:
            ws = wb["Team"]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                data = dict(zip(headers, row))
                member = TeamMember(
                    member_alias=str(data.get("member_alias")),
                    role=str(data.get("role") or ""),
                    # Metrics will be calculated by Risk Engine later
                    projects_in_portfolio=0,
                    open_tasks_assigned=0,
                    blocked_tasks_assigned=0,
                    high_or_critical_open=0,
                    diagnostico_projects=0,
                    proyecto_projects=0,
                    mantenimiento_projects=0,
                )
                await team_repo.create(member)
        logger.info("Seeded Team.")

        # 2. Seed Projects
        if "Projects" in wb.sheetnames:
            ws = wb["Projects"]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                data = dict(zip(headers, row))
                
                business_value = data.get("business_value")
                if business_value is not None:
                    business_value = float(str(business_value).replace(",", ""))

                project = Project(
                    project_code=str(data.get("project_code")),
                    engagement_type=str(data.get("engagement_type")),
                    client_alias=str(data.get("client_alias")),
                    project_name=str(data.get("project_name")),
                    project_type_api=str(data.get("project_type_api")),
                    stage=str(data.get("stage")),
                    status=str(data.get("status", "Activo")),
                    health=str(data.get("health")),
                    owner_alias=str(data.get("owner_alias")),
                    owner_role=str(data.get("owner_role") or ""),
                    start_date=parse_date(data.get("start_date")),
                    target_date=parse_date(data.get("target_date")),
                    business_value=business_value,
                    currency=str(data.get("currency")) if data.get("currency") else None,
                    # Counters will be recalculated
                    open_tasks=0,
                    overdue_tasks=0,
                )
                await project_repo.create(project)
        logger.info("Seeded Projects.")

        # 3. Seed Tasks
        if "Tasks" in wb.sheetnames:
            ws = wb["Tasks"]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]: continue
                data = dict(zip(headers, row))
                
                is_overdue = str(data.get("is_overdue")).lower() in ["si", "yes", "true", "1"]

                task = Task(
                    task_code=str(data.get("task_code")),
                    project_code=str(data.get("project_code")),
                    assignee_alias=str(data.get("assignee_alias")),
                    assignee_role=str(data.get("assignee_role") or ""),
                    priority=str(data.get("priority")),
                    status=str(data.get("status")),
                    due_date=parse_date(data.get("due_date")),
                    is_overdue=is_overdue,
                    dependency=str(data.get("dependency")) if data.get("dependency") else None,
                    title=str(data.get("title")),
                    detail=str(data.get("detail")) if data.get("detail") else None,
                    last_progress=str(data.get("last_progress")) if data.get("last_progress") else None,
                    engagement_type=str(data.get("engagement_type")),
                    client_alias=str(data.get("client_alias")),
                    project_name=str(data.get("project_name")),
                )
                await task_repo.create(task)
        logger.info("Seeded Tasks.")
        
        # 4. Commit everything so Risk Engine has data
        await session.commit()
        
    # 5. Run Risk Engine in a new session block to update metrics
    async with async_session_factory() as session:
        risk_engine = RiskEngineService(session)
        await risk_engine.evaluate_portfolio()
        await session.commit()
        
    logger.info("Seed and metric recalculation completed successfully.")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    asyncio.run(seed_db())
